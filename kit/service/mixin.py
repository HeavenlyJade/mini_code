from abc import ABCMeta, abstractmethod
from contextlib import contextmanager
from typing import List, Optional, Tuple, Type

from openpyxl import load_workbook

from kit.schema.base import ArgSchema

__all__ = ['ExcelImpMixin']


class ExcelImpMixin(metaclass=ABCMeta):
    @property
    @abstractmethod
    def validate_schema(self) -> Type[ArgSchema]:
        """Excel导入校验Schema"""
        ...

    @property
    @abstractmethod
    def sheet_fields(self) -> Tuple[str]:
        """Sheet字段"""
        ...

    @contextmanager
    def load_wb(self, filename):
        wb = load_workbook(filename, read_only=True)
        yield wb
        wb.close()

    def get_items(self, args: dict, attrs: Optional[dict] = None) -> List[dict]:
        schema = self.validate_schema()
        with self.load_wb(args['filename']) as wb:
            ws = wb.active
            items = list()
            for row in ws.iter_rows(min_row=2):
                item = dict()
                for index, cell in enumerate(row):
                    key = self.sheet_fields[index]
                    item[key] = cell.value
                item = schema.load(item)
                if attrs:
                    item.update(attrs)
                items.append(item)
            return items
